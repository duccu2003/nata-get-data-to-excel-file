import re
import logging
import math
from openpyxl import load_workbook
import json
from openpyxl.utils import get_column_letter
logger = logging.getLogger(__name__)

def replace_placeholders_in_sheet(sheet, replacements):
    placeholder_pattern = re.compile(r'\(\d+\)')
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                matches = placeholder_pattern.findall(cell.value)
                if matches:
                    new_value = cell.value
                    for match in matches:
                        placeholder_key = match[1:-1]
                        if placeholder_key in replacements:
                            new_value = new_value.replace(match, str(replacements[placeholder_key]))
                        else:
                            logger.warning(f"No replacement found for placeholder: {match}")
                    cell.value = new_value


def update_dpl_sheet(sheet, dpl_data):
    try:
        containers = json.loads(dpl_data)
    except json.JSONDecodeError:
        raise ValueError("Invalid DPL data format")

    total_row = None
    bulk_in_row = None
    cellmark_row = None
    
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True), 1):
        if row and row[1] and "TOTAL" in str(row[1]):
            total_row = idx
        if row and row[1] and "BULK IN" in str(row[1]):
            bulk_in_row = idx
        if row and row[1] and "Cellmark" in str(row[1]):
            cellmark_row = idx

    if not total_row or not bulk_in_row or not cellmark_row:
        raise ValueError("Could not find summary rows in DPL sheet")

    for row in range(3, total_row):
        sheet[f'B{row}'] = None
        sheet[f'C{row}'] = None
        sheet[f'D{row}'] = None
        sheet[f'E{row}'] = None
        sheet[f'F{row}'] = None

    num_containers = len(containers)
    
    available_rows = total_row - 3
    
    if num_containers > available_rows:
        rows_to_add = num_containers - available_rows
        sheet.insert_rows(3, amount=rows_to_add)
        total_row += rows_to_add
        bulk_in_row += rows_to_add
        cellmark_row += rows_to_add

    for i, container in enumerate(containers):
        row_num = 3 + i
        sheet[f'B{row_num}'] = i + 1
        sheet[f'C{row_num}'] = container.get('contNo', '')
        sheet[f'D{row_num}'] = container.get('sealNo', '')
        sheet[f'E{row_num}'] = container.get('weightNet', 0)
        sheet[f'F{row_num}'] = container.get('weightGross', 0)

    total_net = sum(container.get('weightNet', 0) for container in containers)
    total_gross = sum(container.get('weightGross', 0) for container in containers)
    
    container_weight = 0
    container_weight = total_net / num_containers if num_containers > 0 else 0
    container_weight = math.ceil(container_weight)

    sheet[f'E{total_row}'] = total_net
    sheet[f'F{total_row}'] = total_gross
    sheet[f'B{bulk_in_row}'] = f"BULK IN : {num_containers} X {container_weight} GP CONTAINERS"


def update_csht_sheet(sheet, csht_data):
    try:
        containers = json.loads(csht_data)
    except json.JSONDecodeError:
        raise ValueError("Invalid CSHT data format")

    
    start_row = None
    for idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if row and row[0] and "STT" in str(row[0]):
            start_row = idx + 1
            break
    
    if not start_row:
        raise ValueError("Could not find header row in CSHT sheet")

    num_containers = len(containers)

    total_net = sum(container.get('weightNet', 0) for container in containers)
    total_gross = sum(container.get('weightGross', 0) for container in containers)
    
    container_weight = 0
    container_weight = total_net / num_containers if num_containers > 0 else 0
    container_weight = math.ceil(container_weight)

    for row in range(start_row, sheet.max_row + 1):
        sheet[f'A{row}'] = None
        sheet[f'B{row}'] = None
        sheet[f'C{row}'] = None
        sheet[f'D{row}'] = None
        sheet[f'E{row}'] = None
        sheet[f'F{row}'] = None
        sheet[f'G{row}'] = None

    available_rows = sheet.max_row - start_row + 1
    if num_containers > available_rows:
        rows_to_add = num_containers - available_rows
        sheet.insert_rows(start_row, amount=rows_to_add)

    # Thêm dữ liệu mới
    for i, container in enumerate(containers):
        row_num = start_row + i
        sheet[f'A{row_num}'] = i + 1
        sheet[f'B{row_num}'] = container.get('bookingNumber', '')
        sheet[f'C{row_num}'] = container.get('contNo', '')
        sheet[f'D{row_num}'] = container.get('sealNo', '')
        sheet[f'E{row_num}'] = container_weight  # Loại container (tạm thời ghi sẵn "20")
        sheet[f'F{row_num}'] = "KHO"  # Tính chất container (tạm thời ghi sẵn "KHO")
        sheet[f'G{row_num}'] = ""  # Ghi chú (tạm thời để trống)

    
    