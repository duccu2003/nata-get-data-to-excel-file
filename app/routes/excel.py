import os
import uuid
import logging
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from models.template_data import TemplateData
from utils.excel_utils import replace_placeholders_in_sheet, update_dpl_sheet
from utils.mapping import process_replacements

logger = logging.getLogger(__name__)
router = APIRouter()

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
logger.debug(f"BASE_DIR: {BASE_DIR}")

TEMPLATE_PATH = os.path.join(os.getcwd(), "temp.xlsx")
# TEMPLATE_PATH = r"C:\Users\Acer\Desktop\Nata\Smartwood\apiGetDataToExcel\temp.xlsx"
logger.debug(f"TEMPLATE_PATH: {os.path.abspath(TEMPLATE_PATH)}")

@router.post("/generate-excel/")
async def generate_excel(data: dict):
    try:
        logger.debug(f"Received data: {data}")
        logger.debug(f"Template path: {os.path.abspath(TEMPLATE_PATH)}")
        
        if not os.path.exists(TEMPLATE_PATH):
            raise HTTPException(status_code=404, detail=f"Template file not found at {os.path.abspath(TEMPLATE_PATH)}")

        wb = load_workbook(TEMPLATE_PATH)
        flat_replacements = process_replacements(data)
        dpl_data = data.get('replacements', {}).get('DPL')  # Extract DPL from replacements
        if dpl_data:
            if 'DPL' in wb.sheetnames:
                update_dpl_sheet(wb['DPL'], dpl_data)
            else:
                logger.warning("DPL sheet not found in template, skipping DPL data processing")
        
        for sheet_name in wb.sheetnames:
            if sheet_name != 'DPL':
                sheet = wb[sheet_name]
                replace_placeholders_in_sheet(sheet, flat_replacements)

        os.makedirs(os.path.join(os.getcwd(), "temp"), exist_ok=True)
        output_filename = f"output_{uuid.uuid4()}.xlsx"
        output_path = os.path.join(os.getcwd(), "temp", output_filename)
        wb.save(output_path)
        
        logger.debug(f"File generated at: {output_path}")
        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        logger.error(f"Error generating Excel file: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error generating Excel file: {str(e)}")

